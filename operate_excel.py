import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import os
import time
from openpyxl.styles import PatternFill
from openpyxl.styles import Color, Font, Alignment

import logging
import logging.config
CON_LOG='config\\log.conf'
logging.config.fileConfig(CON_LOG)
logging=logging.getLogger()
def creat_excel(filename=None):
    dt = datetime.now()
    now_time = dt.strftime('%Y_%m_%d_%H_%M_%S')  # 得用下划线，用： 号无法截图保存
    my_path = os.path.abspath(os.getcwd())
    if filename==None:
        filename = my_path + '/Result/result_%s.xlsx' % (now_time)
    else:#为了gui加上的，ui页面中要选中路径，传进来的是路径，不包括文件名
        filename=filename+'/result_%s.xlsx' %(now_time)

    wb=openpyxl.Workbook()
    mysheet=wb.active

    mysheet.merge_cells('A1:F1')
    mysheet.cell(row=1,column=1,value="唤醒成功率测试结果")
    mysheet.row_dimensions[1].height = 25

    # 然后如下设置：
    # 设置表头字体居中
    mycell=mysheet['A1']
    mycell.font=Font(name=u'宋体', bold=True)
    mycell.alignment=Alignment(horizontal='center', vertical='center')

    result_head=['次数','结果','原始日志路径','截取后日志路径','audio文件路径','设备ID']
    for i,item in enumerate(result_head):
        print(i,item)
        mysheet.cell(row=2, column=i + 1, value=item).alignment=Alignment(horizontal='center',vertical='center')
    # mysheet['F2'].font=Font(name=u'宋体',bold=True)
    # mysheet['H2'].font=Font(name=u'宋体',bold=True)

    mysheet.column_dimensions['A'].width = 40
    mysheet.column_dimensions['B'].width = 40
    mysheet.column_dimensions['C'].width = 60
    mysheet.column_dimensions['D'].width = 60
    mysheet.column_dimensions['E'].width = 60
    mysheet.column_dimensions['F'].width = 60


    mysheet.title="测试结果"
    #mysheet.row_dimensions[3].height=25  #设置行高,设置第3行的行高

    wb.save(filename)
    #print(filename)
    logging.info("结果excel保存路径： "+str(filename))
    return filename

def write_excel(filename,num,result,logpath1,logpath2,audiopath,deviceid):
    wb=load_workbook(filename)
    sheet=wb.active
    i=sheet.max_row
    sheet.row_dimensions[i+1].height = 25

    if result==False:
        #填充颜色
        red_fill = PatternFill(fill_type='solid', fgColor="ff441f")
        sheet.cell(row=i + 1, column=1, value=num).fill=red_fill
        sheet.cell(row=i + 1, column=2, value="fail").fill=red_fill
        sheet.cell(row=i + 1, column=3, value=logpath1).fill = red_fill
        sheet.cell(row=i + 1, column=4, value=logpath2).fill = red_fill
        sheet.cell(row=i + 1, column=5, value=audiopath).fill = red_fill
        sheet.cell(row=i + 1, column=6, value=deviceid).fill = red_fill
    else:
        green_file = PatternFill(fill_type='solid', fgColor="16feb5")
        sheet.cell(row=i + 1, column=1, value=num).fill = green_file
        sheet.cell(row=i + 1, column=2, value="pass").fill = green_file
        sheet.cell(row=i + 1, column=6, value=deviceid)

    sheet.cell(row=i + 1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    sheet.cell(row=i + 1, column=2).alignment = Alignment(horizontal='center', vertical='center')
    sheet.cell(row=i + 1, column=6).alignment = Alignment(horizontal='center', vertical='center')
    wb.save(filename)



if __name__=="__main__":
    filename=creat_excel()
    write_excel(filename,1,True,1,1,1)
    write_excel(filename, 2, False,1,1,1)
    write_excel(filename, 3, True,1,1,1)